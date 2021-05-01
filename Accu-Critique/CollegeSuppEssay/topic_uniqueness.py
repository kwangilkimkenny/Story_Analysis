# Chrome 버전을 확인하고 드라이버 버전을 동일하게 해야 함
# 크롬드라이버 다운로드 링크 : https://chromedriver.chromium.org/downloads
# 현재 사용하는 크롬 버전 90.0.4430.85(공식 빌드) (x86_64)
# 적용한 크롬 드라이버 버번은 위와 동일

## 기능 ##
# collegeSupp.py에서 키워드를 추출하여, 본 코드를 활용하여 구글검색 실행하여, 검색결과수를 추출하고
# 결과수를 비교계산하여 토픽의 Uniqueness를 계산한다.

# 만약, 사용자들이 데이터를 검색했다면, 데이터베이스에서 추출한 값을 결과로 가져온다.
# 처음 검색하는 데이터라면 구글검색을 실행하고, 결과값을 동시에 DB에 저장하고 결과도 추력된다.



import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from urllib.parse import ParseResultBytes, quote_plus
from bs4 import BeautifulSoup
from selenium import webdriver

options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
# 혹은 options.add_argument("--disable-gpu")

# UserAgent값을 바꿔줍시다! 서버가 인식하지 못하도록 가상으로 headless 값 추가함ㅠ
options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36")


def get_uniquness(search_result):
    # Dim sum  = 197,000,000
    # Macarons  = 68,100,000
    # Churros   = 24,300,000
    # pasta    = 754,000,000
    # ramen    = 164,000,000
    # udon soba = 30,200,000
    # Tom yam kung = 944,000
    if search_result > 100000000:
        uniqueness_re = 'Common'
        unique_score = 30
    elif search_result <= 100000000 and search_result > 30000000:
        uniqueness_re = 'Unique'
        unique_score = 60
    else: # search_result <= 100000000
        uniqueness_re = 'very unique'
        unique_score = 90
    return uniqueness_re, unique_score


def Go_GoogleSearch(input_word):
        baseUrl = 'https://www.google.com/search?q='

        #plusUrl = input('무엇을 검색할까요? :')
        plusUrl = input_word


        # url = baseUrl + quote_plus(plusUrl)
        url = baseUrl + plusUrl
        # 한글을 사용할 경우 :  quote_plus 적용 - URL에 막 %CE%GD%EC 이런 거 생성해줌

        #office
        #driver = webdriver.Chrome(executable_path= r'./data/chromedriver', chrome_options=options)
        #home
        driver = webdriver.Chrome(executable_path= r'./data/chromedriver_mac_ver_90', chrome_options=options)
        driver.get(url)

        html = driver.page_source
        soup = BeautifulSoup(html, features="html.parser")

        #print(soup.find('div', id='result-stats'))
        get_result = soup.find('div', id='result-stats')

        driver.close()

        result = str(get_result)
        re_ =re.sub("[^()]+$", "", result)
        re__ =re.sub("\([^)]*\)", "", re_) # 괄호안의 문제 제거
        re_d = re.findall("\d+", re__)
        # 검색어로 추출한 결과물을 가지고 topic uniquness 기능을 적용. 검색결과가 평균값(비교 단어로 추정하여 정함)보다 작으면 unique, 크면 ununiqe topis 이다.
        search_re = "".join(re_d)
        #print(search_re)
        input_search_num = int(search_re)
        
        
        # 검색 및 결과데이터 저장기능 (csv로 저장방법)
        # data =[input_word, input_search_num]
        # dataframe = pd.DataFrame(data)
        # dataframe.to_csv("./data/topic_search_result.csv", header=False, index=False)


        # 검색 및 결과데이터 저장기능 (excel로 저장방법) ---> 이 코드를 사용
        # 기존의 저장데이터 불러오기
        wb = openpyxl.load_workbook("./data/topic_search_result.xlsx")

        data =[input_word, input_search_num]
        #wb = openpyxl.Workbook() # 처음 저장할 때 사용
        sheet = wb.active
        sheet.append(data)
        wb.save("./data/topic_search_result.xlsx")

        uniq_result = get_uniquness(input_search_num)

        return uniq_result


# 이것이 1타고 실행되는 함수다!!! 
def google_search_result(input_word):

    # 엑셀로 저장한 데이터의 키워드값을 검색하여, 새로 검색하려는 키워드와 데이터가 존재한다면, 크롤링하지 않도록 기능추가
    # 엑셀데이터를 불러온다. 리스트에 담고, 검색하려는 키워드가 존재하는지 확인한다. 있다면 value 값을 가져온다. 
    # 존재하지 않는다면 크로링기능을 수행한다. 이하는 -- 데이터 검색, 저장 기능 반복됨
    # 데이터가 계속 추가된다.... 끝! 

    # 저장된 엑셀데이터를 불러온다.
    def check_keywd_xlsx_data(input_wd):
        # 데이터 불러오기
        wb = openpyxl.load_workbook("./data/topic_search_result.xlsx")
        load_ws = wb['Sheet']

        # 워크시트의 데이터 모두 불러오기
        all_values = []
        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        #print(all_values)

        # 키워드가 워크시트 데이터에 있는지 확인하고, 있다면 True 와 벨류값 추출하기
        for i in all_values:
            if input_word in i[0]:
                #print(i[1])
                result = [i[0],i[1]]
            else:
                result = None

        return result

    # 검색하려는 키워드가 DB에 있는지 확인
    ck_keywd = check_keywd_xlsx_data(input_word)

    # 검색결과에서 ...
    if ck_keywd == None: # 데이터에 카워드가 없다면 검색을 실행한다.
        result_uniqueness = Go_GoogleSearch(input_word) # 구글검색 실행, 데이터추출, 저장 완료
    else: # 데이터에 키워드가 있은가 그냥 데어값 추력하면 됨
        result_uniqueness = get_uniquness(ck_keywd[1])


    return result_uniqueness



## run ##
result = google_search_result('tesla')
#result = google_search_result('AI')
print('result :' , result)

# result : ('very unique', 90)